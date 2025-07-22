from flask import Flask, render_template_string, request, send_file
import openpyxl
import os
import re
import unicodedata
from openpyxl.styles import PatternFill, Border, Side, Font

# ====== Các hàm chuẩn hóa và xác định ưu tiên (bản gốc ok.py) ======

def no_accent_vn(text):
    if not isinstance(text, str):
        return text
    text = unicodedata.normalize('NFD', text)
    text = re.sub(r'[\u0300-\u036f]', '', text)
    text = re.sub(r'đ', 'd', text)
    text = re.sub(r'Đ', 'D', text)
    return text

def normalize_kw(kw):
    if not isinstance(kw, str): return ""
    kw = no_accent_vn(kw).lower()
    kw = re.sub(r'[\W_]+', ' ', kw)
    kw = kw.strip()
    return kw

RAW_PRIORITY_KEYWORDS = {
    4: [
        "mth", "rung", "chập chờn", "cố định", "chỉnh lại góc", "chỉnh góc", "chỉnh lại cam",
        "chỉnh cam", "bị che", "hồng ngoại", "mic", "micro", "vị trí", "tráo", "đổi", "jack", "lỏng",
        "hư", "bị lỗi", "mờ", "tháo"
    ],
    3: [
        "dữ liệu", "nohdd", "record", "hdd", "norecord"
    ],
    2: [
        "vệ sinh"
    ],
    1: [
        "lỗi màu"
    ],
}

PRIORITY_KEYWORDS = {
    pri: [normalize_kw(x) for x in kw_list]
    for pri, kw_list in RAW_PRIORITY_KEYWORDS.items()
}

def get_priority(text):
    if not text:
        return 0
    text_norm = normalize_kw(text)
    for pri in sorted(PRIORITY_KEYWORDS.keys(), reverse=True):
        for kw in PRIORITY_KEYWORDS[pri]:
            if kw in text_norm:
                return pri
    return 0

# =============== Hàm xử lý file Excel (giống process ok.py) ===============

def process(input_file, input_sheet):
    try:
        wb = openpyxl.load_workbook(input_file)
        if input_sheet not in wb.sheetnames:
            raise Exception("Tên sheet không tồn tại!")
        ws = wb[input_sheet]

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = 'Danh_sach_loi'

        # Copy header (dòng đầu)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            out_ws.cell(row=1, column=col).value = cell.value

        result_rows = []

        # Duyệt từng dòng
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            cell_a = row[0]
            if cell_a.fill.start_color.rgb in ('FFFFFF00', 'FFFF00', '00FFFF00', 'FFFFFF00'):
                # Dòng nền vàng: bỏ qua
                continue

            col_i = row[8].value  # cột I
            col_j = row[9].value  # cột J
            col_k = row[10].value # cột K

            if (col_i and str(col_i).strip()) or (col_j and str(col_j).strip()) or (col_k and str(col_k).strip()):
                # Ít nhất 1 lỗi ở I, J, K
                values = [cell.value for cell in row]
                result_rows.append(values)

                # Nếu có dữ liệu ở cột I, tạo thêm dòng copy với I sang J
                if col_i and str(col_i).strip():
                    new_row = values.copy()
                    new_row[9] = new_row[8]  # J = I
                    result_rows.append(new_row)

        # Ghi dữ liệu ra sheet mới
        for ridx, row in enumerate(result_rows, start=2):
            for cidx, val in enumerate(row, start=1):
                out_ws.cell(row=ridx, column=cidx).value = val

        # Xử lý ưu tiên và tô màu
        for row in out_ws.iter_rows(min_row=2, max_row=out_ws.max_row):
            text = str(row[5].value) if row[5].value else "" # cột F
            prio = get_priority(text)
            fill = None
            if prio == 4 or prio == 2:
                fill = PatternFill(start_color="92D050", fill_type="solid")
            elif prio == 3:
                fill = PatternFill(start_color="FFFF00", fill_type="solid")
            elif prio == 1:
                fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFFFFF", fill_type="solid")
                row[9].fill = PatternFill(start_color="FF0000", fill_type="solid")  # J đỏ nếu không ưu tiên

            row[5].fill = fill  # F
            # Xử lý J trắng hoặc cam nếu là dòng thêm (copy I sang J)
            if row[9].value == row[8].value and row[8].value:  # Nếu J = I và I không rỗng
                row[9].fill = PatternFill(start_color="FFC000", fill_type="solid")  # J cam

            # Ghi thêm cột cuối (ưu tiên)
            prio_str = ""
            if prio == 3 or prio == 4:
                prio_str = "X (ưu tiên 3-4)"
            out_ws.cell(row=row[0].row, column=out_ws.max_column+1).value = prio_str

        # Thêm border cho toàn bộ bảng
        thin = Side(border_style="thin", color="000000")
        for row in out_ws.iter_rows(min_row=1, max_row=out_ws.max_row, max_col=out_ws.max_column):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.font = Font(size=10)

        out_wb.save("ket_qua.xlsx")
        wb.close()
        out_wb.close()
    except Exception as e:
        raise Exception(f"Lỗi xử lý file: {e}")

# ============== Flask App ==============

app = Flask(__name__)

HTML_FORM = '''
<!doctype html>
<html>
<head>
    <title>Sort xe Bình Anh -- Web ver</title>
</head>
<body style="font-family:sans-serif">
    <h2>SORT xe Bình Anh - PhươngĐH</h2>
    <form method="post" enctype="multipart/form-data">
        <label>Chọn file Excel: <input type="file" name="file" accept=".xlsx" required></label><br><br>
        <label>Tên sheet: <input type="text" name="sheet" value="Sheet1" required></label><br><br>
        <button type="submit">Chạy SORT</button>
    </form>
    {% if msg %}<p style="color: red">{{msg}}</p>{% endif %}
</body>
</html>
'''

@app.route("/", methods=["GET", "POST"])
def index():
    msg = ""
    if request.method == "POST":
        file = request.files.get("file")
        sheet_name = request.form.get("sheet")
        if not file or not sheet_name:
            msg = "Vui lòng chọn file và nhập tên sheet!"
            return render_template_string(HTML_FORM, msg=msg)
        try:
            file.save("input.xlsx")
            process("input.xlsx", sheet_name)
            return send_file("ket_qua.xlsx", as_attachment=True)
        except Exception as e:
            msg = f"Lỗi: {e}"
            return render_template_string(HTML_FORM, msg=msg)
    return render_template_string(HTML_FORM, msg=msg)

if __name__ == "__main__":
    app.run(debug=True)
